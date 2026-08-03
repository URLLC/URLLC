from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable


def read_rows(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [dict(row) for row in csv.DictReader(handle)]
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Install Excel support with: pip install -e .[xlsx]") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(values)]
        return [
            {headers[index]: "" if value is None else str(value) for index, value in enumerate(row)}
            for row in values
        ]
    raise ValueError("Input must be .csv or .xlsx")


def write_rows(path: Path, rows: Iterable[dict[str, str]]) -> None:
    records = list(rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    if not records:
        raise ValueError("No rows to write")
    headers = list(records[0].keys())
    if path.suffix.lower() == ".csv":
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers)
            writer.writeheader()
            writer.writerows(records)
        return
    if path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise RuntimeError("Install Excel support with: pip install -e .[xlsx]") from exc
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "cleaned"
        sheet.append(headers)
        for record in records:
            sheet.append([record.get(header, "") for header in headers])
        workbook.save(path)
        return
    raise ValueError("Output must be .csv or .xlsx")

